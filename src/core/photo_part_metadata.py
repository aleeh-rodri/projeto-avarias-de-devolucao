from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.schemas import CHECKLIST_PART_IDS, PART_IDS


VALID_PART_IDS = set(PART_IDS) | set(CHECKLIST_PART_IDS)


def extract_photo_part_code(filename: str) -> str | None:
    """
    Extrai o primeiro número após FOTDEV_.

    Exemplos:
    - FOTDEV_111.jpg -> "111"
    - FOTDEV_164.png -> "164"
    - FOTDEV_111_02.jpg -> "111"
    """
    stem = Path(filename).stem.upper().strip()

    match = re.match(r"^FOTDEV_(\d+)(?:_|$)", stem)
    if not match:
        return None

    return match.group(1)


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _norm_part_id(value: Any) -> str:
    return str(value or "").strip().lower()


class PhotoPartMetadataCache:
    """
    Lê uma planilha que mapeia o id da foto para o part_id canônico.

    Colunas esperadas, com nomes flexíveis:
    - id_foto, id, codigo_foto, photo_code
    - part_id
    - descricao_parte, parte, descricao, item  (opcional)
    """

    ID_COLUMN_ALIASES = {
        "id_foto",
        "id",
        "codigo_foto",
        "código_foto",
        "photo_code",
        "foto_id",
    }

    DESCRIPTION_COLUMN_ALIASES = {
        "descricao_parte",
        "descrição_parte",
        "parte",
        "descricao",
        "descrição",
        "item",
    }

    def __init__(self, xlsx_path: str | Path):
        self.xlsx_path = Path(xlsx_path)
        self._cache = self._load()

    def _find_col(self, header_map: dict[str, int], aliases: set[str]) -> int | None:
        for alias in aliases:
            if alias in header_map:
                return header_map[alias]
        return None

    def _load(self) -> dict[str, dict[str, Any]]:
        if not self.xlsx_path.exists():
            return {}

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        try:
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                return {}

            header_map = {
                _norm_header(header): idx
                for idx, header in enumerate(headers)
                if header is not None
            }

            id_col = self._find_col(header_map, self.ID_COLUMN_ALIASES)
            part_id_col = header_map.get("part_id")
            desc_col = self._find_col(header_map, self.DESCRIPTION_COLUMN_ALIASES)

            if id_col is None:
                raise ValueError(
                    "Planilha de mapeamento sem coluna de id da foto. "
                    "Use uma coluna como: id_foto, id, codigo_foto ou photo_code."
                )

            if part_id_col is None:
                raise ValueError(
                    "Planilha de mapeamento sem coluna 'part_id'."
                )

            out: dict[str, dict[str, Any]] = {}

            for row in rows:
                if row is None:
                    continue

                raw_code = row[id_col] if id_col < len(row) else None
                raw_part_id = row[part_id_col] if part_id_col < len(row) else None

                code = str(raw_code or "").strip()
                part_id = _norm_part_id(raw_part_id)

                if not code or not part_id:
                    continue

                # Se vier 111.0 do Excel, normaliza para 111
                if code.endswith(".0"):
                    code = code[:-2]

                if part_id not in VALID_PART_IDS:
                    raise ValueError(
                        f"part_id inválido na planilha: '{part_id}' para id_foto '{code}'. "
                        f"Cadastre um part_id existente em PART_IDS/CHECKLIST_PART_IDS."
                    )

                description = ""
                if desc_col is not None and desc_col < len(row):
                    description = str(row[desc_col] or "").strip()

                out[code] = {
                    "photo_part_code": code,
                    "part_id": part_id,
                    "description": description,
                }

            return out
        finally:
            wb.close()

    def get(self, photo_part_code: str) -> dict[str, Any] | None:
        code = str(photo_part_code or "").strip()

        if code.endswith(".0"):
            code = code[:-2]

        return self._cache.get(code)