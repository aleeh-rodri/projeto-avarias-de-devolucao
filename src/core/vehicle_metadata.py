from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

# def normalize_wheel_type(valor: str) -> str:
#     v = (valor or "").lower()

#     if "liga leve" in v or "liga" in v:
#         return "liga_leve"

#     if "aco" in v or "aço" in v:
#         return "ferro"

#     return "desconhecido"


# def get_vehicle_wheel_type(placa: str) -> str:
#     query = f"""
#     SELECT
#       attr.valor_atributo
#     FROM `lclz-dados.corporativo_master_data.veiculo`,
#     UNNEST(atributos_modelo) AS attr
#     WHERE attr.nome_atributo = 'WHEEL'
#       AND placa_veiculo = @placa
#     LIMIT 1
#     """

#     job_config = bigquery.QueryJobConfig(
#         query_parameters=[
#             bigquery.ScalarQueryParameter("placa", "STRING", placa)
#         ]
#     )

#     rows = client.query(query, job_config=job_config).result()

#     for row in rows:
#         return normalize_wheel_type(row.valor_atributo)

#     return "desconhecido"




class VehicleMetadataCache:

    def __init__(self, xlsx_path: str | Path):
        self.xlsx_path = Path(xlsx_path)
        self._cache = self._load()

    def _load(self) -> dict[str, dict]:
        if not self.xlsx_path.exists():
            return {}

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        try:
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                return {}

            header_map = {
                str(header).strip().lower(): idx
                for idx, header in enumerate(headers)
                if header is not None
            }

            if "placa" not in header_map or "wheel_type" not in header_map:
                return {}

            out = {}

            for row in rows:
                placa = str(row[header_map["placa"]] or "").strip().upper()

                if not placa:
                    continue

                wheel_raw = None
                if "wheel_raw" in header_map:
                    wheel_raw = row[header_map["wheel_raw"]]

                out[placa] = {
                    "wheel_raw": wheel_raw,
                    "wheel_type": str(row[header_map["wheel_type"]] or "desconhecido").strip(),
                    "template_type": str(row[header_map.get("template_type", "")] or "GF").strip().upper()
                    if "template_type" in header_map
                    else "GF",
                }

            return out
        finally:
            wb.close()

    def get_wheel_type(self, placa: str) -> str:
        placa_norm = str(placa).strip().upper()

        item = self._cache.get(placa_norm)

        if not item:
            return "desconhecido"

        return str(item.get("wheel_type") or "desconhecido")

    def get_vehicle_metadata(self, placa: str) -> dict:
        placa_norm = str(placa).strip().upper()
        item = self._cache.get(placa_norm)

        if not item:
            return {"wheel_type": "desconhecido",
                    "template_type": "GF",
                    }

        return dict(item)

