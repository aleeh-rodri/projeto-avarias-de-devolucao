import json
import os
import re
import warnings
from pathlib import Path
from openpyxl import load_workbook
from typing import Any
from openpyxl.drawing.image import Image as OpenpyxlImage
from copy import copy
from PIL import Image as PILImage
from PIL import ImageOps
from core.pdf_utils import extract_checklist_text, extract_reldev_avaria_items

class ExcelAgent:
    def __init__(self, template_path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template não encontrado: {template_path}")

    def _clear_orcamento_itens(self, ws, *, start_row: int = 24, end_row: int = 180) -> None:
        """Limpa a tabela de itens no template (descrição/qtd/valor).

        O template vem com linhas de exemplo (ex.: 'avaria 1/2' com 320) e também pode
        ficar com sobra de linhas quando o caso atual tem menos itens.
        Para não herdar esses valores, zeramos as colunas D/E/F em um range seguro.

        Importante: não limpamos B/C para preservar as fórmulas do template (numeração/tipo).
        """
        # Nota: `Worksheet.cell(..., value=None)` NÃO limpa a célula (openpyxl só atribui
        # quando o `value` é diferente de None). Por isso, limpamos via `.value = None`.
        for r in range(int(start_row), int(end_row) + 1):
            ws.cell(row=r, column=4).value = None  # Descrição
            ws.cell(row=r, column=5).value = None  # Qtd.
            ws.cell(row=r, column=6).value = None  # Valor

    def _clear_unused_detalhamento_blocks(self, ws, *, used_blocks: int, first_block_row: int = 16, block_rows: int = 16) -> None:
        """Limpa os blocos não usados na aba Detalhamento_Avarias.

        O template traz 2 blocos base (Avaria 1 e Avaria 2) com fórmulas apontando
        para Orçamento_Padrão. Se o caso tiver menos itens, esses blocos podem ficar
        aparecendo (ex.: número 2 e valor). Aqui removemos os conteúdos do cabeçalho
        do bloco para "sumir" a seção.
        """
        try:
            max_blocks = max(0, (int(ws.max_row) - int(first_block_row) + 1) // int(block_rows))
        except Exception:
            return

        for block_idx in range(int(used_blocks), int(max_blocks)):
            row = int(first_block_row) + (block_idx * int(block_rows))
            # Células com conteúdo/fórmula no cabeçalho do bloco
            ws.cell(row=row, column=2).value = None  # número da avaria
            ws.cell(row=row, column=3).value = None  # descrição (pode ter fórmula)
            ws.cell(row=row, column=10).value = None  # label "Valor" do bloco
            ws.cell(row=row, column=11).value = None  # valor (pode ter fórmula)

    def _copy_cell_style(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def _copy_row_formatting(self, ws, source_row, target_row, max_col=12):
        for col in range(1, max_col + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            self._copy_cell_style(source_cell, target_cell)
        
        # Copia a altura da linha se estiver definida
        if source_row in ws.row_dimensions:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def _copy_block_formatting(self, ws, source_start_row: int, target_start_row: int, block_rows: int = 16, max_col: int = 12) -> None:
        """Copia a formatação de um bloco (múltiplas linhas) mantendo merges.

        O template de Detalhamento_Avarias usa um bloco por avaria.
        Copiar só 1 linha não replica corretamente as células de fundo (ex.: verde).
        """

        # 1) Copiar estilos linha a linha
        for i in range(block_rows):
            self._copy_row_formatting(ws, source_start_row + i, target_start_row + i, max_col=max_col)

        # 2) Replicar merges que existirem dentro do bloco
        #    (necessário para preservar áreas de layout do template)
        try:
            ranges = list(ws.merged_cells.ranges)
        except Exception:
            ranges = []

        row_delta = target_start_row - source_start_row
        for r in ranges:
            min_col, min_row, max_col_r, max_row = r.bounds

            if not (source_start_row <= min_row and max_row <= source_start_row + block_rows - 1):
                continue

            new_min_row = min_row + row_delta
            new_max_row = max_row + row_delta
            merge_str = f"{ws.cell(row=new_min_row, column=min_col).coordinate}:{ws.cell(row=new_max_row, column=max_col_r).coordinate}"
            try:
                ws.merge_cells(merge_str)
            except ValueError:
                # merge já existe
                pass

    def _strip_known_prefixes(self, descricao: str) -> str:
        """Remove prefixos padronizados já existentes para não duplicar rótulos no Excel."""
        d = (descricao or "").strip()
        # Normaliza separadores mais comuns do projeto
        for pfx in (
            "CHECKLIST —",
            "CHECKLIST -",
            "CHECKLIST:",
            "REVISAR —",
            "REVISAR -",
            "REVISAR:",
            "AGENTE_IA-avaria —",
            "AGENTE_IA-avaria -",
            "AGENTE_IA-avaria:",
            "AGENTE_IA —",
            "AGENTE_IA -",
            "AGENTE_IA:",
        ):
            if d.upper().startswith(pfx.upper()):
                d = d[len(pfx):].strip()
                break
        return d

    def _format_descricao_origem(self, descricao: str, *, origem: str | None) -> str:
        base = self._strip_known_prefixes(descricao)

        if origem == "checklist":
            return f"CHECKLIST - {base}" if base else "CHECKLIST -"

        if origem == "agente_ia":
            # conforme pedido: diferenciar claramente quando veio do nosso agente
            return f"AGENTE_IA-avaria - {base}" if base else "AGENTE_IA-avaria -"

        return descricao

    def _parts_with_sem_dano_from_laudo(self, laudo_data: dict[str, Any]) -> set[str]:
        """Retorna part_ids onde algum perito concluiu `sem_dano` usando imagens dessa peça.

        Usado para suprimir fallback de checklist no Excel quando já houve análise visual
        apontando ausência de dano (evita relatório misleading com falso positivo do checklist).
        """
        out: set[str] = set()
        peritos = laudo_data.get("peritos", {}) if isinstance(laudo_data, dict) else {}
        if not isinstance(peritos, dict):
            return out

        for _perito_name, pdata in peritos.items():
            if not isinstance(pdata, dict):
                continue
            resultado = pdata.get("resultado")
            if not isinstance(resultado, dict):
                continue

            nivel = str(resultado.get("nivel_dano") or "").strip().lower()
            if nivel != "sem_dano":
                continue

            servicos = resultado.get("servicos")
            if isinstance(servicos, list) and servicos:
                # Se tem serviços, não é um "sem dano" puro.
                continue

            imagens_usadas = pdata.get("imagens_usadas")
            if not isinstance(imagens_usadas, list):
                continue

            for img in imagens_usadas:
                if not isinstance(img, dict):
                    continue
                part_id = str(img.get("part_id") or "").strip()
                if part_id:
                    out.add(part_id)

        return out

    def extract_info_from_pdf(self, pdf_path):
        text = extract_checklist_text(str(pdf_path))
        info = {
            "cliente": "xxxx",
            "contrato": "xxxx",
            "modelo": "xxxx",
            "cor": "xxxx",
            "placa": "xxxx"
        }
        
        if not text:
            return info

        # Regex para Modelo, Cor e Placa (baseado no padrão observado no PDF)
        # Exemplo: ARGO DRIVE 1.0 6V FLEX 4P C/AR BRANCO RUY5J95
        padrao_carro = re.search(r"Informa\w\w\wes do carroModeloCorPlaca\s*(.*?)\s*(BRANCO|PRETO|CINZA|PRATA|VERMELHO|VERDE|AZUL)\s*([A-Z]{3}[0-9][A-Z0-9][0-9]{2})", text)
        if padrao_carro:
            info["modelo"] = padrao_carro.group(1).strip()
            info["cor"] = padrao_carro.group(2).strip()
            info["placa"] = padrao_carro.group(3).strip()
        else:
            # Fallback patterns
            m_match = re.search(r"Modelo\s*(.*?)\s*Cor", text)
            if m_match: info["modelo"] = m_match.group(1).strip()
            
            c_match = re.search(r"Cor\s*(.*?)\s*Placa", text)
            if c_match: info["cor"] = c_match.group(1).strip()
            
            p_match = re.search(r"Placa\s*([A-Z]{3}[0-9][A-Z0-9][0-9]{2})", text)
            if p_match: info["placa"] = p_match.group(1).strip()

        # Cliente e Contrato: Deixados como 'xxxx' conforme solicitado
        
        return info

    def _prepare_image_for_excel(self, image_path: str, temp_img_path: str, base_height: int = 300) -> None:
        """Normaliza orientação (EXIF) e redimensiona para colar no Excel.

        Observação: a correção de orientação aqui é baseada em EXIF Orientation.
        Se a imagem foi salva "girada no pixel" e sem EXIF, não há como inferir
        automaticamente a rotação correta de forma determinística.
        """
        with PILImage.open(image_path) as pil_img:
            # Corrige rotação baseada em EXIF (quando existir)
            pil_img = ImageOps.exif_transpose(pil_img)

            # Redimensiona mantendo aspecto
            w_percent = (base_height / float(pil_img.size[1]))
            w_size = int((float(pil_img.size[0]) * float(w_percent)))
            pil_img = pil_img.resize((w_size, base_height), PILImage.Resampling.LANCZOS)

            pil_img.save(temp_img_path)

    def _resolve_photo_path(
        self,
        photo_path: str | None,
        *,
        case_id: str | None,
        laudo_path: str | Path,
    ) -> str | None:
        """Tenta resolver caminhos quebrados/antigos de foto para um path existente.

        Estratégia (ordem):
        1) Se o path original existir, usa ele.
        2) Se houver case_id, tenta localizar pelo basename em:
           - <BASE_DIR>/input/<case_id>/Fotos
           - <BASE_DIR>/input/<case_id>/fotos
        3) Tenta também na pasta do laudo (output do case): <laudo_dir>/<basename>
        """
        if not photo_path:
            return None

        try:
            p = Path(str(photo_path))
        except Exception:
            return None

        if p.exists():
            return str(p)

        name = p.name
        if not name:
            return None

        laudo_dir = Path(laudo_path).parent
        cand_out = laudo_dir / name
        if cand_out.exists():
            return str(cand_out)

        if case_id:
            base_dir = Path(__file__).resolve().parents[2]
            for folder in (base_dir / "input" / str(case_id) / "Fotos", base_dir / "input" / str(case_id) / "fotos"):
                cand = folder / name
                if cand.exists():
                    return str(cand)

        return None

    def _load_triage_index(self, laudo_path: str | Path) -> dict[str, dict[str, Any]]:
        """Carrega um índice image_id -> metadados da triagem (part_id, confidence, checklist_damage_reported).

        Usa o triage.json na mesma pasta do laudo.
        """
        laudo_path = Path(laudo_path)
        triage_path = laudo_path.parent / "triage.json"
        if not triage_path.exists():
            return {}

        try:
            triage = json.loads(triage_path.read_text(encoding="utf-8"))
            images = triage.get("images", [])
            if not isinstance(images, list):
                return {}

            idx: dict[str, dict[str, Any]] = {}
            for img in images:
                if not isinstance(img, dict):
                    continue
                image_id = str(img.get("image_id") or "").strip()
                if not image_id:
                    continue
                idx[image_id] = {
                    "part_id": str(img.get("part_id") or "").strip(),
                    "confidence": float(img.get("confidence") or 0.0),
                    "checklist_damage_reported": img.get("checklist_damage_reported"),
                }
            return idx
        except Exception:
            return {}

    def _load_triage_raw(self, laudo_path: str | Path) -> dict[str, Any] | None:
        """Carrega o triage.json na mesma pasta do laudo (raw)."""
        laudo_path = Path(laudo_path)
        triage_path = laudo_path.parent / "triage.json"
        if not triage_path.exists():
            return None
        try:
            triage = json.loads(triage_path.read_text(encoding="utf-8"))
            return triage if isinstance(triage, dict) else None
        except Exception:
            return None

    def _compute_checklist_fallback_lines(
        self,
        *,
        laudo_path: str | Path,
        case_id: str | None,
        triage_raw: dict[str, Any] | None,
        triage_index: dict[str, dict[str, Any]],
        current_servicos: list[dict[str, Any]],
        checklist_part_ids: set[str] | None,
        checklist_avaria_items: list[Any] | None = None,
    ) -> list[dict[str, Any]]:
        """Cria linhas de revisão humana para peças marcadas no checklist que não viraram cobrança.

        Isso é um fallback "no Excel" (compatibilidade), para casos em que o laudo
        não tenha `cobrancas_checklist_fallback`.
        """
        # Preferência: se temos o PDF do checklist (RELDEV) e conseguimos extrair
        # as peças marcadas como AVARIA, usamos isso como fonte de verdade.
        best_photo_by_part: dict[str, str] = {}
        best_conf_by_part: dict[str, float] = {}

        checklist_entries: list[dict[str, Any]] = []
        seen_known_parts: set[str] = set()

        if checklist_avaria_items:
            for idx, item in enumerate(checklist_avaria_items):
                part_id = str(getattr(item, "part_id", "") or "").strip()
                raw_label = str(getattr(item, "raw_label", "") or "").strip()
                page = getattr(item, "page", None)

                if part_id:
                    if part_id in seen_known_parts:
                        continue
                    seen_known_parts.add(part_id)
                    checklist_entries.append(
                        {
                            "key": f"part:{part_id}",
                            "part_id": part_id,
                            "label": part_id,
                            "raw_label": raw_label,
                        }
                    )
                    continue

                if not raw_label:
                    continue
                checklist_entries.append(
                    {
                        "key": f"raw:{page or 0}:{idx}",
                        "part_id": "",
                        "label": raw_label,
                        "raw_label": raw_label,
                    }
                )

        elif checklist_part_ids:
            for part_id in sorted(set(checklist_part_ids or [])):
                checklist_entries.append(
                    {
                        "key": f"part:{part_id}",
                        "part_id": part_id,
                        "label": part_id,
                        "raw_label": part_id,
                    }
                )

        # Se não temos checklist_part_ids, cai para o comportamento antigo baseado no triage.json.
        if not checklist_entries:
            if not triage_raw or not isinstance(triage_raw.get("images"), list):
                return []

            checklist_parts: set[str] = set()
            for img in triage_raw.get("images", []) or []:
                if not isinstance(img, dict):
                    continue
                if img.get("checklist_damage_reported") is not True:
                    continue
                part_id = str(img.get("part_id") or "").strip()
                if not part_id:
                    continue
                checklist_parts.add(part_id)

            for part_id in sorted(checklist_parts):
                checklist_entries.append(
                    {
                        "key": f"part:{part_id}",
                        "part_id": part_id,
                        "label": part_id,
                        "raw_label": part_id,
                    }
                )

        if not checklist_entries:
            return []

        checklist_parts = {
            str(entry.get("part_id") or "").strip()
            for entry in checklist_entries
            if entry.get("part_id")
        }

        # Escolhe a melhor foto por peça (se houver triage)
        if triage_raw and isinstance(triage_raw.get("images"), list):
            for img in triage_raw.get("images", []) or []:
                if not isinstance(img, dict):
                    continue
                part_id = str(img.get("part_id") or "").strip()
                if not part_id or part_id not in checklist_parts:
                    continue
                photo_path = str(img.get("photo_path") or "").strip()
                try:
                    conf = float(img.get("confidence") or 0.0)
                except Exception:
                    conf = 0.0
                if part_id not in best_conf_by_part or conf > best_conf_by_part.get(part_id, 0.0):
                    best_conf_by_part[part_id] = conf
                    best_photo_by_part[part_id] = photo_path

        # Peças que já viraram cobrança (a partir do triage_meta nas linhas atuais)
        charged_parts: set[str] = set()
        for s in current_servicos or []:
            explicit_part_id = str(s.get("part_id") or "").strip() if isinstance(s, dict) else ""
            if explicit_part_id:
                charged_parts.add(explicit_part_id)
                continue

            tm = s.get("triage_meta") if isinstance(s, dict) else None
            part_id = str((tm or {}).get("part_id") or "").strip()
            if part_id:
                charged_parts.add(part_id)

        missing_entries = [
            entry
            for entry in checklist_entries
            if not entry.get("part_id") or str(entry.get("part_id") or "").strip() not in charged_parts
        ]
        if not missing_entries:
            return []

        out: list[dict[str, Any]] = []
        for entry in missing_entries:
            part_id = str(entry.get("part_id") or "").strip()
            label = str(entry.get("label") or entry.get("raw_label") or part_id or "item sem part_id").strip()
            raw_photo = best_photo_by_part.get(part_id) or None
            resolved = self._resolve_photo_path(raw_photo, case_id=case_id, laudo_path=laudo_path)
            triage_meta = self._triage_meta_for_photo(triage_index, resolved or raw_photo)
            out.append(
                {
                    "descricao": f"Avaria reportada na peca '{label}' (REVISAR)",
                    "valor": 0,
                    "fotos": [resolved] if resolved else ([raw_photo] if raw_photo else []),
                    "triage_meta": triage_meta,
                    "force_include": True,
                    "origin": "checklist_fallback_excel",
                    "part_id": part_id,
                }
            )

        return out

    def _triage_meta_for_photo(self, triage_index: dict[str, dict[str, Any]], photo_path: str | None) -> dict[str, Any] | None:
        if not triage_index or not photo_path:
            return None
        try:
            image_id = Path(photo_path).stem  # nome do arquivo sem extensão (bate com triage.image_id)
        except Exception:
            return None
        return triage_index.get(image_id)

    def _should_include_by_checklist(
        self,
        triage_meta: dict[str, Any] | None,
        *,
        checklist_only: bool,
        extra_conf_threshold: float,
        checklist_part_ids: set[str] | None,
    ) -> bool:
        """Regra:
        - Se checklist_only=False: sempre inclui.
        - Se não houver sinal do checklist (meta ausente): inclui (fallback compatível).
        - Se checklist_damage_reported=True: inclui.
        - Caso contrário: só inclui se confidence > extra_conf_threshold.
        """
        if not checklist_only:
            return True
        if not triage_meta:
            return True

        # Se temos a fonte de verdade do checklist (via PDF), usa ela.
        if checklist_part_ids:
            part_id = str(triage_meta.get("part_id") or "").strip()
            if part_id and part_id in checklist_part_ids:
                return True
        else:
            # Comportamento legado
            if triage_meta.get("checklist_damage_reported") is True:
                return True
        try:
            conf = float(triage_meta.get("confidence") or 0.0)
        except Exception:
            conf = 0.0
        return conf > float(extra_conf_threshold)

    def generate_report(self, laudo_path, output_path, pdf_path=None, *, checklist_only: bool = True, extra_conf_threshold: float = 0.9):
        """Gera o relatório Excel a partir do laudo.json.

        Layout atual esperado na aba Orçamento_Padrão:
        - C18: cliente (preencher "xxx")
        - C19: contrato (preencher "xxx")
        - C20: placa
        - C21: modelo
        - E19: total geral (fórmula do template)
        - Avarias: B32:F65
        - Total avarias: F67 (fórmula do template)
        - Peças/Mão de obra: B73:F106
        - Total peças: F109 (fórmula do template)

        Observação: a coluna C das tabelas já possui fórmula no template.
        Por isso, o código só preenche B, D, E e F.
        """
        with open(laudo_path, 'r', encoding='utf-8') as f:
            laudo_data = json.load(f)

        triage_index = self._load_triage_index(laudo_path)
        has_triage = bool(triage_index)
        triage_raw = self._load_triage_raw(laudo_path)

        # Fonte de verdade do checklist (quando o PDF do RELDEV estiver disponível)
        checklist_part_ids: set[str] | None = None
        checklist_avaria_items: list[Any] | None = None
        if pdf_path and os.path.exists(pdf_path):
            try:
                checklist_avaria_items = extract_reldev_avaria_items(pdf_path)
                checklist_part_ids = {
                    str(item.part_id).strip()
                    for item in checklist_avaria_items
                    if getattr(item, "part_id", None)
                }
            except Exception:
                checklist_part_ids = None

        # O template pode conter extensões de formatação condicional que o openpyxl
        # não suporta (ele remove ao carregar). Isso é esperado e não deve poluir o log.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"Conditional Formatting extension is not supported and will be removed",
                category=UserWarning,
                module=r"openpyxl\..*",
            )
            wb = load_workbook(self.template_path)

        # Extrair dados do PDF se existir.
        # Cliente/contrato ficam fixos como "xxx" por enquanto, conforme definição atual.
        pdf_info = {
            "cliente": "xxx",
            "contrato": "xxx",
            "modelo": "xxxx",
            "cor": "xxxx",
            "placa": laudo_data.get("case_id", "xxxx")
        }
        if pdf_path and os.path.exists(pdf_path):
            extracted_info = self.extract_info_from_pdf(pdf_path)
            pdf_info["modelo"] = extracted_info.get("modelo") or pdf_info["modelo"]
            pdf_info["cor"] = extracted_info.get("cor") or pdf_info["cor"]
            pdf_info["placa"] = extracted_info.get("placa") or pdf_info["placa"]

        # =========================
        # 1. Aba Orçamento_Padrão
        # =========================
        ws_orcamento = wb["Orçamento_Padrão"]

        AVARIAS_START_ROW = 32
        AVARIAS_END_ROW = 65
        PECAS_START_ROW = 73
        PECAS_END_ROW = 106

        def _clear_budget_section(start_row: int, end_row: int) -> None:
            """Limpa B/D/E/F preservando C, pois C tem fórmula no template."""
            for r in range(int(start_row), int(end_row) + 1):
                ws_orcamento.cell(row=r, column=2).value = None  # Num. item
                ws_orcamento.cell(row=r, column=4).value = None  # Descrição
                ws_orcamento.cell(row=r, column=5).value = None  # Qtd.
                ws_orcamento.cell(row=r, column=6).value = None  # Valor

        _clear_budget_section(AVARIAS_START_ROW, AVARIAS_END_ROW)
        _clear_budget_section(PECAS_START_ROW, PECAS_END_ROW)

        # Cabeçalho do novo template
        ws_orcamento["C18"] = "xxx"
        ws_orcamento["C19"] = "xxx"
        ws_orcamento["C20"] = pdf_info["placa"]
        ws_orcamento["C21"] = pdf_info["modelo"]

        # Mantém E19/F67/F109 com as fórmulas do template.
        # Não sobrescrever fórmulas aqui.

        all_servicos: list[dict[str, Any]] = []
        all_pecas_a_cotar: list[dict[str, Any]] = []

        for perito_name, perito_data in laudo_data.get("peritos", {}).items():
            if not isinstance(perito_data, dict):
                continue

            resultado = perito_data.get("resultado", {})
            if not isinstance(resultado, dict):
                continue

            perito_force_include = bool(resultado.get("force_include") is True)
            perito_needs_human_review = bool(resultado.get("needs_human_review") is True)
            resultado_origin = str(resultado.get("origin") or "").strip()
            # Heurística de origem: só rotular como AGENTE_IA quando for linha "manual/revisão"
            # (ex.: force_include/needs_human_review). Serviços normais do LPU ficam sem rótulo.
            perito_origem = (
                None
                if resultado_origin == "checklist_chave_reserva_visual"
                else (
                    "checklist"
                    if resultado_origin.startswith("checklist_")
                    else ("agente_ia" if (perito_force_include or perito_needs_human_review) else None)
                )
            )

            # Preferir fotos por item (mais preciso) quando houver breakdown em "itens".
            itens = resultado.get("itens")
            if isinstance(itens, list) and itens:
                for it in itens:
                    if not isinstance(it, dict):
                        continue

                    fotos_it = it.get("fotos_analisadas", [])
                    servicos_it = it.get("servicos", [])
                    if not isinstance(servicos_it, list):
                        continue

                    for svc_idx, svc in enumerate(servicos_it):
                        if not isinstance(svc, dict):
                            continue

                        # IMPORTANT: o template suporta 1 foto por linha.
                        # Se um item tem múltiplos serviços e também múltiplas fotos,
                        # pareia cada serviço com a foto do mesmo índice.
                        fotos_line: list[str] = []
                        if isinstance(fotos_it, list) and fotos_it:
                            if svc_idx < len(fotos_it):
                                fotos_line = [fotos_it[svc_idx]]
                            else:
                                fotos_line = [fotos_it[0]]
                        else:
                            fotos_res = resultado.get("fotos_analisadas", [])
                            if isinstance(fotos_res, list) and fotos_res:
                                fotos_line = [fotos_res[0]]

                        triage_meta = self._triage_meta_for_photo(triage_index, fotos_line[0] if fotos_line else None)
                        all_servicos.append({
                            "descricao": self._format_descricao_origem(svc.get("descricao", ""), origem=perito_origem),
                            "valor": svc.get("preco", 0),
                            "qtd": 1,
                            "fotos": fotos_line,
                            "triage_meta": triage_meta,
                            "force_include": perito_force_include or bool(it.get("force_include") is True),
                        })
            else:
                servicos = resultado.get("servicos", [])
                if isinstance(servicos, list):
                    for svc in servicos:
                        if not isinstance(svc, dict):
                            continue
                        fotos_res = resultado.get("fotos_analisadas", [])
                        fotos_line = [fotos_res[0]] if isinstance(fotos_res, list) and fotos_res else []
                        triage_meta = self._triage_meta_for_photo(triage_index, fotos_line[0] if fotos_line else None)
                        all_servicos.append({
                            "descricao": self._format_descricao_origem(svc.get("descricao", ""), origem=perito_origem),
                            "valor": svc.get("preco", 0),
                            "qtd": 1,
                            "fotos": fotos_line,
                            "triage_meta": triage_meta,
                            "force_include": perito_force_include,
                        })

            pecas = resultado.get("pecas_a_cotar", [])
            if isinstance(pecas, list):
                for p in pecas:
                    if not isinstance(p, dict):
                        continue

                    desc = p.get("descricao", "")
                    obs = p.get("observacao", "")
                    qtd = p.get("quantidade", 1)
                    if not desc:
                        continue

                    full_desc = f"PEÇA (cotação manual): {desc}"
                    if obs:
                        full_desc += f" — {obs}"

                    fotos_p = p.get("fotos_analisadas")
                    if not isinstance(fotos_p, list) or not fotos_p:
                        fotos_p = (resultado.get("fotos_analisadas", []) or [])
                    triage_meta = self._triage_meta_for_photo(
                        triage_index,
                        fotos_p[0] if isinstance(fotos_p, list) and fotos_p else None,
                    )
                    all_pecas_a_cotar.append({
                        "descricao": full_desc,
                        "valor": 0,  # valor da peça será preenchido manualmente
                        "qtd": qtd if isinstance(qtd, int) else 1,
                        "fotos": fotos_p,
                        "triage_meta": triage_meta,
                        "force_include": perito_force_include,
                    })

        # =========================================================
        # Fallback do checklist: quando o checklist marcou avaria, mas nenhum serviço foi gerado.
        # Esses itens devem aparecer no Excel com sinalização de revisão humana.
        # =========================================================
        fallback = laudo_data.get("cobrancas_checklist_fallback", [])
        if isinstance(fallback, list) and fallback:
            for fb in fallback:
                if not isinstance(fb, dict):
                    continue

                fotos_fb = fb.get("fotos", [])
                fotos_line: list[str] = []
                if isinstance(fotos_fb, list) and fotos_fb:
                    fotos_line = [fotos_fb[0]] if isinstance(fotos_fb[0], str) else []

                triage_meta = self._triage_meta_for_photo(triage_index, fotos_line[0] if fotos_line else None)
                all_servicos.append({
                    "descricao": self._format_descricao_origem(
                        fb.get("descricao", "Avaria reportada (REVISAR)"),
                        origem="checklist",
                    ),
                    "valor": fb.get("valor", 0),
                    "qtd": 1,
                    "fotos": fotos_line,
                    "triage_meta": triage_meta,
                    "force_include": True,
                    "part_id": str(fb.get("part_id") or "").strip(),
                })

        # Complementa fallbacks usando o checklist PDF como fonte principal.
        # Se não houver PDF/checklist_part_ids, usa o fallback antigo baseado na triagem.
        should_compute_fallback = bool(checklist_avaria_items) or bool(checklist_part_ids) or ((not isinstance(fallback, list) or not fallback) and has_triage)
        if should_compute_fallback:
            computed_fb = self._compute_checklist_fallback_lines(
                laudo_path=laudo_path,
                case_id=pdf_info.get("placa"),
                triage_raw=triage_raw,
                triage_index=triage_index,
                current_servicos=all_servicos,
                checklist_part_ids=checklist_part_ids,
                checklist_avaria_items=checklist_avaria_items,
            )
            # Garantir prefixo CHECKLIST no modo compatibilidade também
            for s in computed_fb:
                if not isinstance(s, dict):
                    continue
                s["descricao"] = self._format_descricao_origem(str(s.get("descricao") or ""), origem="checklist")
                if "qtd" not in s:
                    s["qtd"] = 1
                all_servicos.append(s)

        # =========================================================
        # FILTRO: cobrar somente o que estiver no checklist.
        # Exceção: quando confiança da triagem > extra_conf_threshold.
        # =========================================================
        if has_triage:
            all_servicos = [
                svc
                for svc in all_servicos
                if (
                    (not checklist_only)
                    or (svc.get("force_include") is True)
                    or (svc.get("triage_meta") is not None and self._should_include_by_checklist(
                        svc.get("triage_meta"),
                        checklist_only=checklist_only,
                        extra_conf_threshold=extra_conf_threshold,
                        checklist_part_ids=checklist_part_ids,
                    ))
                )
            ]
            all_pecas_a_cotar = [
                p
                for p in all_pecas_a_cotar
                if (
                    (not checklist_only)
                    or (p.get("force_include") is True)
                    or (p.get("triage_meta") is not None and self._should_include_by_checklist(
                        p.get("triage_meta"),
                        checklist_only=checklist_only,
                        extra_conf_threshold=extra_conf_threshold,
                        checklist_part_ids=checklist_part_ids,
                    ))
                )
            ]

        def _is_mao_de_obra(item: dict[str, Any]) -> bool:
            desc = str(item.get("descricao") or "").strip().lower()
            desc_norm = (
                desc.replace("ã", "a")
                .replace("á", "a")
                .replace("à", "a")
                .replace("â", "a")
                .replace("é", "e")
                .replace("ê", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ô", "o")
                .replace("ú", "u")
                .replace("ç", "c")
            )
            return "mao de obra" in desc_norm or "m.o" in desc_norm or "mo " in f"{desc_norm} "

        # Nova regra:
        # - Avarias: serviços normais, exceto mão de obra.
        # - Peças a cotar: peças a cotar + qualquer serviço identificado como mão de obra.
        avarias_items: list[dict[str, Any]] = []
        pecas_mao_obra_items: list[dict[str, Any]] = []

        for svc in all_servicos:
            if _is_mao_de_obra(svc):
                pecas_mao_obra_items.append(svc)
            else:
                avarias_items.append(svc)

        pecas_mao_obra_items.extend(all_pecas_a_cotar)

        def _write_budget_rows(items: list[dict[str, Any]], *, start_row: int, end_row: int, item_num_start: int) -> int:
            row = int(start_row)
            item_num = int(item_num_start)
            overflow = 0

            for item in items:
                if row > int(end_row):
                    overflow += 1
                    continue

                ws_orcamento.cell(row=row, column=2, value=item_num)
                # Coluna C preservada: fórmula do template.
                ws_orcamento.cell(row=row, column=4, value=item.get("descricao", ""))
                ws_orcamento.cell(row=row, column=5, value=item.get("qtd", 1))
                ws_orcamento.cell(row=row, column=6, value=item.get("valor", 0))

                row += 1
                item_num += 1

            if overflow:
                print(
                    f"Aviso: {overflow} item(ns) excederam o limite de linhas do template "
                    f"({start_row}:{end_row}) e não foram escritos no orçamento."
                )

            return item_num

        next_item_num = _write_budget_rows(
            avarias_items,
            start_row=AVARIAS_START_ROW,
            end_row=AVARIAS_END_ROW,
            item_num_start=1,
        )
        _write_budget_rows(
            pecas_mao_obra_items,
            start_row=PECAS_START_ROW,
            end_row=PECAS_END_ROW,
            item_num_start=next_item_num,
        )

        # =========================
        # 2. Aba Detalhamento_Avarias
        # =========================
        ws_detalhe = wb["Detalhamento_Avarias"]
        ws_detalhe["C7"] = "xxx"
        ws_detalhe["C8"] = "xxx"
        # ws_detalhe["C10"] = pdf_info["placa"]
        # ws_detalhe["C11"] = pdf_info["modelo"]
        # ws_detalhe["C12"] = pdf_info["cor"]

        DETAIL_FIRST_ITEM_ROW = 14
        DETAIL_BLOCK_ROWS = 16
        DETAIL_IMAGE_OFFSET = 2

        detail_items = list(avarias_items) + list(pecas_mao_obra_items)

        for idx, svc in enumerate(detail_items):
            header_row = DETAIL_FIRST_ITEM_ROW + (idx * DETAIL_BLOCK_ROWS)
            image_row = header_row + DETAIL_IMAGE_OFFSET

            raw_photo = svc["fotos"][0] if (
                isinstance(svc.get("fotos"), list) and svc["fotos"]
            ) else None

            img_path = self._resolve_photo_path(
                raw_photo,
                case_id=pdf_info.get("placa"),
                laudo_path=laudo_path,
            )

            if img_path and os.path.exists(img_path):
                try:
                    temp_img_path = f"temp_img_{idx}_{pdf_info['placa']}.png"
                    self._prepare_image_for_excel(img_path, temp_img_path, base_height=300)

                    img = OpenpyxlImage(temp_img_path)
                    ws_detalhe.add_image(img, f"B{image_row}")
                except Exception as e:
                    print(f"Erro ao inserir imagem: {e}")

        # Se o template tiver blocos pré-criados (ex.: Avaria 2), limpa os que não foram usados
        # para evitar que apareçam valores do template.
        # self._clear_unused_detalhamento_blocks(
        #     ws_detalhe,
        #     used_blocks=len(detail_items),
        #     first_block_row=16,
        #     block_rows=16,
        # )

        wb.save(output_path)

        # Limpeza
        for idx in range(len(detail_items)):
            temp_path = f"temp_img_{idx}_{pdf_info['placa']}.png"
            if os.path.exists(temp_path):
                os.remove(temp_path)

        print(f"Relatório Excel gerado: {output_path}")
        return output_path
