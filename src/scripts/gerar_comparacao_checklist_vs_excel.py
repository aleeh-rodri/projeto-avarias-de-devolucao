from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

# =========================
# RESOLUÇÃO DE PATH
# =========================
BASE_DIR = Path(__file__).resolve().parents[2]
SRC_DIR = BASE_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from core.pdf_utils import extract_checklist_text  # noqa: E402
from core.input_resolver import resolve_checklist_pdf  # noqa: E402
from core.pdf_utils import extract_reldev_rows  # noqa: E402


# =========================
# MODELOS
# =========================
@dataclass(frozen=True)
class ChecklistAvaria:
    raw: str
    part_key: str


@dataclass(frozen=True)
class ExcelCobranca:
    descricao: str
    valor: float | str
    part_key: str


# =========================
# NORMALIZAÇÃO
# =========================
_RE_SPACES = re.compile(r"\s+")


def _norm(s: str) -> str:
    return _RE_SPACES.sub(" ", (s or "").strip().lower())


def _dedup_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        if x in seen:
            continue
        seen.add(x)
        out.append(x)
    return out


def map_text_to_part_key(text: str) -> str:
    """Mapeia um texto (do checklist ou de uma linha do Excel) para uma chave canônica.

    Retorna "__unknown__" quando não for possível.
    """
    t = _norm(text)

    # Para-choques
    if "para-choque" in t or "parachoque" in t:
        if "dianteir" in t:
            return "parachoque_dianteiro"
        if "traseir" in t:
            return "parachoque_traseiro"
        # Se não tiver posição, deixa genérico
        return "parachoque"

    # Portas
    if "porta" in t:
        if "dianteir" in t:
            if "direit" in t:
                return "porta_dianteira_direita"
            if "esquerd" in t:
                return "porta_dianteira_esquerda"
            return "porta_dianteira"
        if "traseir" in t:
            if "direit" in t:
                return "porta_traseira_direita"
            if "esquerd" in t:
                return "porta_traseira_esquerda"
            return "porta_traseira"
        return "porta"

    # Para-lamas
    if "paralama" in t or "para-lama" in t:
        # No checklist/RELDEV, pode vir apenas 'Para-Lama Esquerdo/Direito'.
        # Por padrão, consideramos dianteiro.
        if "dianteir" not in t and "traseir" not in t:
            if "direit" in t:
                return "paralama_dianteiro_direito"
            if "esquerd" in t:
                return "paralama_dianteiro_esquerdo"

        if "dianteir" in t:
            if "direit" in t:
                return "paralama_dianteiro_direito"
            if "esquerd" in t:
                return "paralama_dianteiro_esquerdo"
            return "paralama_dianteiro"
        if "traseir" in t:
            if "direit" in t:
                return "paralama_traseiro_direito"
            if "esquerd" in t:
                return "paralama_traseiro_esquerdo"
            return "paralama_traseiro"
        return "paralama"

    # Rodas
    if "roda" in t:
        if "dianteir" in t:
            if "direit" in t:
                return "roda_dianteira_direita"
            if "esquerd" in t:
                return "roda_dianteira_esquerda"
            return "roda_dianteira"
        if "traseir" in t:
            if "direit" in t:
                return "roda_traseira_direita"
            if "esquerd" in t:
                return "roda_traseira_esquerda"
            return "roda_traseira"
        return "roda"

    # Vidros
    if "para-brisa" in t or "parabrisa" in t or "pára-brisa" in t:
        return "parabrisa"
    if "vidro" in t and "trase" in t:
        return "vidro_traseiro"

    # Capô / Teto
    if "capo" in t or "capô" in t:
        return "capo"
    if "teto" in t:
        return "teto"

    # Pneus / rodas / calotas
    if "calota" in t or "calotas" in t:
        return "calota"

    # Retrovisores
    if "retrovisor" in t:
        if "direit" in t:
            return "retrovisor_direito"
        if "esquerd" in t:
            return "retrovisor_esquerdo"
        return "retrovisor"

    # Laterais (genérico do checklist)
    if "lateral" in t:
        if "direit" in t:
            return "lateral_direita"
        if "esquerd" in t:
            return "lateral_esquerda"
        return "lateral"

    return "__unknown__"


# =========================
# CHECKLIST
# =========================
def extract_avarias_from_checklist_text(checklist_text: str) -> list[ChecklistAvaria]:
    """Extrai itens marcados como 'Avaria' do texto do Checkfacil.

    Heurística: busca padrões comuns e mapeia para part_key.
    Não tenta reconstruir a tabela inteira (o PDF vem com whitespace colapsado).
    """
    t = " ".join((checklist_text or "").split())

    raw_hits: list[str] = []

    # Para-choque: o PDF costuma trazer "Traseira Para-Choque Avaria" e "Dianteira Do Carro Para-Choque Avaria"
    if re.search(r"Traseira\s*Para-Choque\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Traseira / Para-Choque")
    if re.search(r"Dianteira(\s+Do\s+Carro)?\s*Para-Choque\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Dianteira / Para-Choque")

    # Capô / Teto / Para-brisa
    # Observação: no PDF, às vezes vem colado (ex.: "CarroCapôAvaria"), por isso \s*.
    if re.search(r"Cap[ôo]\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Capô")
    if re.search(r"Teto\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Teto")
    if re.search(r"Para-?Brisa\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Para-brisa")

    # Portas
    for m in re.finditer(
        r"Porta\s+(Dianteira|Traseira)\s+(Direita|Esquerda)\s*Avaria",
        t,
        flags=re.IGNORECASE,
    ):
        raw_hits.append(f"Porta {m.group(1)} {m.group(2)}")

    # Rodas
    for m in re.finditer(
        r"Roda\s+(Dianteira|Traseira)\s+(Direita|Esquerda)\s*Avaria",
        t,
        flags=re.IGNORECASE,
    ):
        raw_hits.append(f"Roda {m.group(1)} {m.group(2)}")

    # Retrovisor
    for m in re.finditer(
        r"Retrovisor\s+(Direito|Esquerdo)\s*Avaria",
        t,
        flags=re.IGNORECASE,
    ):
        raw_hits.append(f"Retrovisor {m.group(1)}")

    # Para-lama
    for m in re.finditer(
        r"Para-?lama\s+(Dianteiro|Traseiro)\s+(Direito|Esquerdo)\s*Avaria",
        t,
        flags=re.IGNORECASE,
    ):
        raw_hits.append(f"Paralama {m.group(1)} {m.group(2)}")

    # Lateral (genérico)
    if re.search(r"Lateral\s+Direita\s*Lateral\s+Direita\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Lateral Direita")
    if re.search(r"Lateral\s+Esquerda\s*Lateral\s+Esquerda\s*Avaria", t, flags=re.IGNORECASE):
        raw_hits.append("Lateral Esquerda")

    raw_hits = _dedup_preserve_order([_RE_SPACES.sub(" ", h).strip() for h in raw_hits if h.strip()])

    out: list[ChecklistAvaria] = []
    for h in raw_hits:
        part_key = map_text_to_part_key(h)
        out.append(ChecklistAvaria(raw=h, part_key=part_key))
    return out


def extract_checklist_avarias(placa: str) -> tuple[list[ChecklistAvaria], str | None]:
    """Tenta extrair o texto do checklist via PDF; se não existir, usa triage.json."""
    checklist_pdf = resolve_checklist_pdf(BASE_DIR / "input" / placa)
    if checklist_pdf and checklist_pdf.exists():
        # Preferência: extração determinística da tabela DESCRIÇÃO/ITEM/REGISTRO do RELDEV.
        try:
            rows = extract_reldev_rows(checklist_pdf)
        except Exception:
            rows = []

        if rows:
            avarias: list[ChecklistAvaria] = []
            for r in rows:
                if _norm(r.registro) != "avaria":
                    continue
                raw = " ".join([x for x in [r.descricao, r.item] if x]).strip()
                part_key = map_text_to_part_key(raw)
                avarias.append(ChecklistAvaria(raw=raw, part_key=part_key))

            if avarias:
                return (avarias, str(checklist_pdf))

        # Fallback: tenta heurística no texto colapsado.
        text = extract_checklist_text(str(checklist_pdf))
        avarias = extract_avarias_from_checklist_text(text)
        if avarias:
            return (avarias, str(checklist_pdf))

    triage_path = BASE_DIR / "output" / placa / "triage.json"
    if triage_path.exists():
        try:
            triage = json.loads(triage_path.read_text(encoding="utf-8"))
            # 1) Preferência: sinal por foto (mais confiável do que regex em texto colapsado)
            images = triage.get("images")
            if isinstance(images, list) and images:
                part_ids: list[str] = []
                for img in images:
                    if not isinstance(img, dict):
                        continue
                    if img.get("checklist_damage_reported") is not True:
                        continue
                    part_id = str(img.get("part_id") or "").strip()
                    if part_id:
                        part_ids.append(part_id)

                # Dedup preservando ordem
                part_ids = _dedup_preserve_order(part_ids)
                if part_ids:
                    out = [ChecklistAvaria(raw=pid, part_key=map_text_to_part_key(pid)) for pid in part_ids]
                    return (out, str(triage_path))

            # 2) Fallback final: tenta heurística no checklist_summary
            text = triage.get("checklist_summary") or ""
            return (extract_avarias_from_checklist_text(text), str(triage_path))
        except Exception:
            pass

    return ([], None)


# =========================
# EXCEL
# =========================
def _to_float_or_none(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").strip())
    except Exception:
        return None


def extract_cobrancas_from_excel(xlsx_path: Path) -> list[ExcelCobranca]:
    """Extrai cobranças (linhas com valor > 0) da aba Orçamento_Padrão."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "Orçamento_Padrão" not in wb.sheetnames:
        # fallback: tenta primeira aba
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb["Orçamento_Padrão"]

    out: list[ExcelCobranca] = []

    # No template atual, a tabela começa em linha 24:
    # col 4 = descrição, col 6 = valor
    row = 24
    empty_streak = 0
    while row <= ws.max_row and empty_streak < 20:
        desc = ws.cell(row=row, column=4).value
        val = ws.cell(row=row, column=6).value

        desc_str = (str(desc).strip() if desc is not None else "")
        if not desc_str:
            empty_streak += 1
            row += 1
            continue
        empty_streak = 0

        # Ignora cabeçalho de seção de peças
        if _norm(desc_str).startswith("peças a cotar"):
            row += 1
            continue

        val_num = _to_float_or_none(val)
        if val_num is not None and val_num > 0:
            part_key = map_text_to_part_key(desc_str)
            out.append(ExcelCobranca(descricao=desc_str, valor=round(val_num, 2), part_key=part_key))
        else:
            # Se for 'Sob consulta' ou algo não numérico, ainda pode ser relevante.
            if isinstance(val, str) and _norm(val) in {"sob consulta", "a cotar"}:
                part_key = map_text_to_part_key(desc_str)
                out.append(ExcelCobranca(descricao=desc_str, valor=str(val).strip(), part_key=part_key))

        row += 1

    return out


# =========================
# MARKDOWN
# =========================
def build_markdown(
    placa: str,
    checklist: list[ChecklistAvaria],
    cobrancas: list[ExcelCobranca],
    checklist_source: str | None,
    excel_path: str | None,
    excel_read_error: str | None = None,
) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    checklist_keys = [c.part_key for c in checklist]
    cobrancas_keys = [c.part_key for c in cobrancas]

    checklist_set = set([k for k in checklist_keys if k != "__unknown__"])
    cobrancas_set = set([k for k in cobrancas_keys if k != "__unknown__"])

    matched = sorted(checklist_set.intersection(cobrancas_set))
    only_checklist = sorted(checklist_set.difference(cobrancas_set))
    only_excel = sorted(cobrancas_set.difference(checklist_set))

    def fmt_key(k: str) -> str:
        return k

    lines: list[str] = []
    lines.append(f"# Comparação Checklist vs Excel — {placa}")
    lines.append("")
    lines.append(f"Gerado em: {now}")
    if checklist_source:
        lines.append(f"Fonte checklist: `{checklist_source}`")
    if excel_path:
        lines.append(f"Fonte excel: `{excel_path}`")
    if excel_read_error:
        lines.append(f"Aviso: não foi possível ler o Excel ({excel_read_error})")
    lines.append("")

    lines.append("## Resumo")
    lines.append("")
    lines.append(f"- Checklist (avarias detectadas): {len(checklist)}")
    lines.append(f"- Excel (linhas cobradas): {len([c for c in cobrancas if isinstance(c.valor, (int, float))])}")
    lines.append(f"- Peças em comum (normalizadas): {len(matched)}")
    lines.append(f"- No checklist e não cobradas: {len(only_checklist)}")
    lines.append(f"- Cobradas e não no checklist: {len(only_excel)}")
    lines.append("")

    lines.append("## Checklist — Avarias marcadas")
    lines.append("")
    if checklist:
        for c in checklist:
            lines.append(f"- {c.raw}  →  `{c.part_key}`")
    else:
        lines.append("- (nenhuma avaria encontrada / não foi possível extrair)")
    lines.append("")

    lines.append("## Excel — Linhas cobradas")
    lines.append("")
    if cobrancas:
        for c in cobrancas:
            lines.append(f"- R$ {c.valor} — {c.descricao}  →  `{c.part_key}`")
    else:
        lines.append("- (nenhuma cobrança encontrada / excel não encontrado)")
    lines.append("")

    lines.append("## Comparação (por peça normalizada)")
    lines.append("")

    lines.append("### Em comum")
    lines.append("")
    if matched:
        for k in matched:
            lines.append(f"- `{fmt_key(k)}`")
    else:
        lines.append("- (nenhuma)")
    lines.append("")

    lines.append("### No checklist e não cobradas")
    lines.append("")
    if only_checklist:
        for k in only_checklist:
            lines.append(f"- `{fmt_key(k)}`")
    else:
        lines.append("- (nenhuma)")
    lines.append("")

    lines.append("### Cobradas e não no checklist")
    lines.append("")
    if only_excel:
        for k in only_excel:
            lines.append(f"- `{fmt_key(k)}`")
    else:
        lines.append("- (nenhuma)")
    lines.append("")

    lines.append("## Observações / Itens não mapeados")
    lines.append("")
    unknown_checklist = [c for c in checklist if c.part_key == "__unknown__"]
    unknown_excel = [c for c in cobrancas if c.part_key == "__unknown__"]

    if not unknown_checklist and not unknown_excel:
        lines.append("- (nenhum)")
        lines.append("")
        return "\n".join(lines)

    if unknown_checklist:
        lines.append("### Checklist (sem mapeamento)")
        lines.append("")
        for c in unknown_checklist:
            lines.append(f"- {c.raw}")
        lines.append("")

    if unknown_excel:
        lines.append("### Excel (sem mapeamento)")
        lines.append("")
        for c in unknown_excel:
            lines.append(f"- {c.descricao} — {c.valor}")
        lines.append("")

    return "\n".join(lines)


# =========================
# EXECUÇÃO
# =========================
def _listar_placas_em_output(output_dir: Path) -> list[str]:
    if not output_dir.exists() or not output_dir.is_dir():
        return []
    placas: list[str] = []
    for child in output_dir.iterdir():
        if child.is_dir() and child.name and not child.name.startswith("."):
            placas.append(child.name.strip())
    placas.sort()
    return placas


def _normalizar_placas(args: Iterable[str]) -> list[str]:
    placas_args = [a.strip() for a in args if a.strip()]
    if placas_args:
        return placas_args
    return _listar_placas_em_output(BASE_DIR / "output")


def gerar_para_placa(placa: str) -> tuple[bool, str]:
    output_dir = BASE_DIR / "output" / placa
    if not output_dir.exists():
        return (False, f"output/{placa} não existe")

    checklist, checklist_source = extract_checklist_avarias(placa)

    xlsx_path = output_dir / f"{placa}.xlsx"
    cobrancas: list[ExcelCobranca] = []
    excel_source: str | None = None
    excel_read_error: str | None = None
    if xlsx_path.exists():
        excel_source = str(xlsx_path)
        try:
            cobrancas = extract_cobrancas_from_excel(xlsx_path)
        except PermissionError as e:
            excel_read_error = f"PermissionError: {e}"
            cobrancas = []
        except Exception as e:
            excel_read_error = f"{type(e).__name__}: {e}"
            cobrancas = []

    md = build_markdown(
        placa=placa,
        checklist=checklist,
        cobrancas=cobrancas,
        checklist_source=checklist_source,
        excel_path=excel_source,
        excel_read_error=excel_read_error,
    )

    out_md = output_dir / "comparacao_checklist_vs_excel.md"
    out_md.write_text(md, encoding="utf-8")
    return (True, str(out_md))


def main() -> int:
    placas = _normalizar_placas(sys.argv[1:])
    if not placas:
        print("Nenhuma placa encontrada (output/ vazio) e nenhuma placa informada.")
        return 2

    ok = 0
    fail = 0
    for idx, placa in enumerate(placas, start=1):
        success, msg = gerar_para_placa(placa)
        status = "OK" if success else "FAIL"
        print(f"[{idx}/{len(placas)}] {status} {placa}: {msg}")
        if success:
            ok += 1
        else:
            fail += 1

    print(f"\nResumo: OK={ok} | FAIL={fail}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
